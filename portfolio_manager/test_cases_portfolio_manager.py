"""
Portfolio Manager - Test Cases Generator

This script generates a comprehensive Excel file with test cases for the 
Portfolio Manager agent that can be used by testing teams for manual testing.
"""

import pandas as pd
from datetime import datetime

def generate_test_cases():
    """Generate comprehensive test cases for Portfolio Manager."""
    
    test_cases = []
    
    # ==================== CATEGORY 1: API HEALTH & INFO ENDPOINTS ====================
    
    test_cases.append({
        "Test Case #": "TC001",
        "Category": "Health Check",
        "Test Case Description": "Verify /health endpoint returns healthy status",
        "Test Type": "Functional",
        "Preconditions": "Server is running on port 8000",
        "Test Steps": "1. Send GET request to http://localhost:8000/health\n2. Check response status code\n3. Verify response body contains status field",
        "Test Data": "N/A",
        "Expected Output": "Status: 200 OK\nResponse body: {'status': 'healthy', 'service': 'portfolio_manager', 'version': '2.0.0', 'agent': 'autonomous_agent'}",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Critical - Basic health check"
    })
    
    test_cases.append({
        "Test Case #": "TC002",
        "Category": "API Info",
        "Test Case Description": "Verify /api/info endpoint returns agent information",
        "Test Type": "Functional",
        "Preconditions": "Server is running",
        "Test Steps": "1. Send GET request to http://localhost:8000/api/info\n2. Verify response contains name, version, agent_type\n3. Check pipeline array has 7 agents",
        "Test Data": "N/A",
        "Expected Output": "Status: 200 OK\nResponse contains: name, version, agent_type='autonomous', pipeline with 7 items, endpoints array",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Verifies agent metadata"
    })
    
    test_cases.append({
        "Test Case #": "TC003",
        "Category": "A2A Protocol",
        "Test Case Description": "Verify agent card endpoint for A2A discovery",
        "Test Type": "Functional",
        "Preconditions": "Server is running",
        "Test Steps": "1. Send GET request to http://localhost:8000/.well-known/agent-card.json\n2. Verify AgentCard structure\n3. Check supportedInterfaces contains /message:send",
        "Test Data": "N/A",
        "Expected Output": "Status: 200 OK\nResponse contains: name='Portfolio Manager', version='2.0.0', supportedInterfaces with /message:send endpoint",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Required for A2A protocol compatibility"
    })
    
    test_cases.append({
        "Test Case #": "TC004",
        "Category": "UI Access",
        "Test Case Description": "Verify web UI loads successfully",
        "Test Type": "UI",
        "Preconditions": "Server is running, ui/index.html exists",
        "Test Steps": "1. Open browser to http://localhost:8000/\n2. Verify page loads without errors\n3. Check for HTML content and UI elements",
        "Test Data": "N/A",
        "Expected Output": "Status: 200 OK\nHTML page loads with Portfolio Manager interface",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Test encoding fix for Windows (UTF-8)"
    })
    
    # ==================== CATEGORY 2: PORTFOLIO GENERATION - VALID INPUTS ====================
    
    test_cases.append({
        "Test Case #": "TC005",
        "Category": "Portfolio Generation",
        "Test Case Description": "Generate portfolio with conservative risk profile",
        "Test Type": "Functional",
        "Preconditions": "Valid API keys in .env, Server running",
        "Test Steps": "1. Send POST to /api/generate-portfolio\n2. Provide conservative profile data\n3. Wait for response (may take 30-60 seconds)\n4. Verify success=true and all sections present",
        "Test Data": '{"capital": 10000, "goal": "preserve_capital", "horizon": "3_5_years", "max_loss": 5, "experience": "beginner", "income_stability": "stable"}',
        "Expected Output": "Status: 200 OK\nsuccess=true\nprofile with risk_score 1-3\nmacro_outlook present\ntop_sectors present\nselected_stocks present\nportfolio with conservative allocations\nperformance_report present\nbacktest_results present\nfinal_report present",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Risk score should be 1-3 (Conservative)"
    })
    
    test_cases.append({
        "Test Case #": "TC006",
        "Category": "Portfolio Generation",
        "Test Case Description": "Generate portfolio with moderate risk profile",
        "Test Type": "Functional",
        "Preconditions": "Valid API keys, Server running",
        "Test Steps": "1. Send POST to /api/generate-portfolio\n2. Provide moderate profile data\n3. Verify balanced allocation (60-70% stocks, 30-40% bonds/stable)",
        "Test Data": '{"capital": 25000, "goal": "balanced_growth", "horizon": "5_10_years", "max_loss": 15, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "Status: 200 OK\nsuccess=true\nrisk_score 5-6 (Moderate)\nportfolio with balanced allocations\nstocks in 4-6 different sectors\ntotal allocation = 100%\nSharpe ratio present in performance_report",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Most common risk profile"
    })
    
    test_cases.append({
        "Test Case #": "TC007",
        "Category": "Portfolio Generation",
        "Test Case Description": "Generate portfolio with aggressive risk profile",
        "Test Type": "Functional",
        "Preconditions": "Valid API keys, Server running",
        "Test Steps": "1. Send POST to /api/generate-portfolio\n2. Provide aggressive profile data\n3. Verify high equity allocation (85-100%)",
        "Test Data": '{"capital": 50000, "goal": "aggressive_growth", "horizon": "10_plus_years", "max_loss": 30, "experience": "advanced", "income_stability": "very_stable"}',
        "Expected Output": "Status: 200 OK\nsuccess=true\nrisk_score 8-10 (Aggressive)\nportfolio with 85-100% stocks\ngrowth-oriented sectors (Technology, Healthcare)\nhigher expected volatility in performance_report",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "High-risk tolerance investor"
    })
    
    test_cases.append({
        "Test Case #": "TC008",
        "Category": "Portfolio Generation",
        "Test Case Description": "Generate portfolio with minimum capital",
        "Test Type": "Boundary",
        "Preconditions": "Valid API keys, Server running",
        "Test Steps": "1. Send POST with capital=100 (minimum)\n2. Verify portfolio generation succeeds\n3. Check allocations are realistic for small capital",
        "Test Data": '{"capital": 100, "goal": "balanced_growth", "horizon": "3_5_years", "max_loss": 15, "experience": "beginner", "income_stability": "stable"}',
        "Expected Output": "Status: 200 OK\nsuccess=true\nportfolio generated with allocations\ntotal_value close to 100\nfractional shares or limited positions",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Edge case - minimum investment"
    })
    
    test_cases.append({
        "Test Case #": "TC009",
        "Category": "Portfolio Generation",
        "Test Case Description": "Generate portfolio with large capital",
        "Test Type": "Boundary",
        "Preconditions": "Valid API keys, Server running",
        "Test Steps": "1. Send POST with capital=1000000 (large)\n2. Verify portfolio scales appropriately\n3. Check diversification across more positions",
        "Test Data": '{"capital": 1000000, "goal": "balanced_growth", "horizon": "5_10_years", "max_loss": 20, "experience": "expert", "income_stability": "very_stable"}',
        "Expected Output": "Status: 200 OK\nsuccess=true\nportfolio with 8-15 positions\nwell-diversified across sectors\ntotal_value close to 1000000",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "High net worth investor scenario"
    })
    
    test_cases.append({
        "Test Case #": "TC010",
        "Category": "Portfolio Generation",
        "Test Case Description": "Generate portfolio with short time horizon",
        "Test Type": "Functional",
        "Preconditions": "Valid API keys, Server running",
        "Test Steps": "1. Send POST with horizon='less_than_1_year'\n2. Verify conservative allocation\n3. Check for low-volatility stocks or bonds",
        "Test Data": '{"capital": 15000, "goal": "preserve_capital", "horizon": "less_than_1_year", "max_loss": 5, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "Status: 200 OK\nsuccess=true\nrisk_score 2-4\nportfolio with defensive stocks or bonds\nlower expected returns\nlower volatility",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Short-term investment scenario"
    })
    
    test_cases.append({
        "Test Case #": "TC011",
        "Category": "Portfolio Generation",
        "Test Case Description": "Generate portfolio with long time horizon",
        "Test Type": "Functional",
        "Preconditions": "Valid API keys, Server running",
        "Test Steps": "1. Send POST with horizon='10_plus_years'\n2. Verify higher risk tolerance in allocation\n3. Check for growth stocks",
        "Test Data": '{"capital": 30000, "goal": "aggressive_growth", "horizon": "10_plus_years", "max_loss": 25, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "Status: 200 OK\nsuccess=true\nrisk_score 7-9\nportfolio with growth stocks\nhigher allocation to equities\nlong-term growth potential",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Long-term retirement scenario"
    })
    
    # ==================== CATEGORY 3: INPUT VALIDATION & ERROR HANDLING ====================
    
    test_cases.append({
        "Test Case #": "TC012",
        "Category": "Input Validation",
        "Test Case Description": "Reject portfolio request with missing required field (capital)",
        "Test Type": "Negative",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST to /api/generate-portfolio without capital field\n2. Verify validation error returned",
        "Test Data": '{"goal": "balanced_growth", "horizon": "3_5_years", "max_loss": 15}',
        "Expected Output": "Status: 422 Unprocessable Entity\nValidation error indicating capital is required",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Required field validation"
    })
    
    test_cases.append({
        "Test Case #": "TC013",
        "Category": "Input Validation",
        "Test Case Description": "Reject capital below minimum (< 100)",
        "Test Type": "Negative",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with capital=50\n2. Verify validation error",
        "Test Data": '{"capital": 50, "goal": "balanced_growth", "horizon": "3_5_years", "max_loss": 15, "experience": "beginner", "income_stability": "stable"}',
        "Expected Output": "Status: 422 Unprocessable Entity\nValidation error: capital must be >= 100",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Minimum capital boundary test"
    })
    
    test_cases.append({
        "Test Case #": "TC014",
        "Category": "Input Validation",
        "Test Case Description": "Reject max_loss below minimum (< 5)",
        "Test Type": "Negative",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with max_loss=3\n2. Verify validation error",
        "Test Data": '{"capital": 10000, "goal": "balanced_growth", "horizon": "3_5_years", "max_loss": 3, "experience": "beginner", "income_stability": "stable"}',
        "Expected Output": "Status: 422 Unprocessable Entity\nValidation error: max_loss must be >= 5",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Risk tolerance minimum validation"
    })
    
    test_cases.append({
        "Test Case #": "TC015",
        "Category": "Input Validation",
        "Test Case Description": "Reject max_loss above maximum (> 50)",
        "Test Type": "Negative",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with max_loss=60\n2. Verify validation error",
        "Test Data": '{"capital": 10000, "goal": "aggressive_growth", "horizon": "10_plus_years", "max_loss": 60, "experience": "expert", "income_stability": "stable"}',
        "Expected Output": "Status: 422 Unprocessable Entity\nValidation error: max_loss must be <= 50",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Risk tolerance maximum validation"
    })
    
    test_cases.append({
        "Test Case #": "TC016",
        "Category": "Input Validation",
        "Test Case Description": "Handle invalid goal value",
        "Test Type": "Negative",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with invalid goal='invalid_goal'\n2. Check if gracefully handled or defaults applied",
        "Test Data": '{"capital": 10000, "goal": "invalid_goal", "horizon": "3_5_years", "max_loss": 15, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "Status: 422 or graceful handling with default to balanced_growth",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Invalid enum value handling"
    })
    
    test_cases.append({
        "Test Case #": "TC017",
        "Category": "Input Validation",
        "Test Case Description": "Handle invalid horizon value",
        "Test Type": "Negative",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with invalid horizon='2_years'\n2. Check error handling",
        "Test Data": '{"capital": 10000, "goal": "balanced_growth", "horizon": "2_years", "max_loss": 15, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "Status: 422 or graceful handling with default",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Invalid enum value for time horizon"
    })
    
    test_cases.append({
        "Test Case #": "TC018",
        "Category": "Input Validation",
        "Test Case Description": "Handle malformed JSON in request body",
        "Test Type": "Negative",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with invalid JSON syntax\n2. Verify 400 or 422 error",
        "Test Data": '{"capital": 10000, "goal": "balanced_growth", INVALID JSON',
        "Expected Output": "Status: 400 Bad Request or 422\nError message about invalid JSON",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Malformed request handling"
    })
    
    # ==================== CATEGORY 4: RISK SCORE CALCULATION ====================
    
    test_cases.append({
        "Test Case #": "TC019",
        "Category": "Risk Calculation",
        "Test Case Description": "Verify risk score calculation for ultra-conservative profile",
        "Test Type": "Functional",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with all conservative parameters\n2. Verify risk_score in response is 1-2\n3. Check risk_category='Ultra Conservative' or 'Very Conservative'",
        "Test Data": '{"capital": 10000, "goal": "preserve_capital", "horizon": "less_than_1_year", "max_loss": 5, "experience": "none", "income_stability": "stable"}',
        "Expected Output": "risk_score: 1-2\nrisk_category: Ultra Conservative or Very Conservative",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Minimum risk score verification"
    })
    
    test_cases.append({
        "Test Case #": "TC020",
        "Category": "Risk Calculation",
        "Test Case Description": "Verify risk score for maximum risk profile",
        "Test Type": "Functional",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with all aggressive parameters\n2. Verify risk_score=9-10\n3. Check risk_category contains 'Aggressive'",
        "Test Data": '{"capital": 50000, "goal": "aggressive_growth", "horizon": "10_plus_years", "max_loss": 50, "experience": "expert", "income_stability": "very_stable"}',
        "Expected Output": "risk_score: 9-10\nrisk_category: Very Aggressive or Maximum Risk",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Maximum risk score verification"
    })
    
    test_cases.append({
        "Test Case #": "TC021",
        "Category": "Risk Calculation",
        "Test Case Description": "Verify risk score for balanced moderate profile",
        "Test Type": "Functional",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST with moderate parameters\n2. Verify risk_score=5-6\n3. Check risk_category='Moderate'",
        "Test Data": '{"capital": 20000, "goal": "balanced_growth", "horizon": "3_5_years", "max_loss": 15, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "risk_score: 5-6\nrisk_category: Moderate or Moderately Aggressive",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Middle-range risk score"
    })
    
    # ==================== CATEGORY 5: OUTPUT VALIDATION ====================
    
    test_cases.append({
        "Test Case #": "TC022",
        "Category": "Output Validation",
        "Test Case Description": "Verify macro_outlook contains required economic indicators",
        "Test Type": "Functional",
        "Preconditions": "Valid portfolio generation request completed",
        "Test Steps": "1. Generate portfolio\n2. Inspect macro_outlook object\n3. Verify contains: GDP, inflation, unemployment, interest_rates",
        "Test Data": "Use TC005 request",
        "Expected Output": "macro_outlook object contains:\n- GDP data or growth rate\n- Inflation rate or CPI\n- Unemployment rate\n- Interest rates or Fed funds rate\n- Market outlook or sentiment",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Macro agent output validation"
    })
    
    test_cases.append({
        "Test Case #": "TC023",
        "Category": "Output Validation",
        "Test Case Description": "Verify top_sectors contains sector recommendations",
        "Test Type": "Functional",
        "Preconditions": "Valid portfolio generation completed",
        "Test Steps": "1. Generate portfolio\n2. Inspect top_sectors object\n3. Verify contains 3-5 sectors with rationale",
        "Test Data": "Use TC006 request",
        "Expected Output": "top_sectors contains:\n- Array or object with 3-5 sectors\n- Each sector has name and reasoning\n- Sectors appropriate for economic conditions",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Sector agent output validation"
    })
    
    test_cases.append({
        "Test Case #": "TC024",
        "Category": "Output Validation",
        "Test Case Description": "Verify selected_stocks contains stock picks",
        "Test Type": "Functional",
        "Preconditions": "Valid portfolio generation completed",
        "Test Steps": "1. Generate portfolio\n2. Inspect selected_stocks object\n3. Verify contains 5-15 stocks with ticker symbols",
        "Test Data": "Use TC006 request",
        "Expected Output": "selected_stocks contains:\n- Array of 5-15 stocks\n- Each stock has: ticker, company name, sector\n- Fundamental metrics (P/E, revenue growth)\n- Technical indicators (RSI, MACD)",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Stock selection agent output"
    })
    
    test_cases.append({
        "Test Case #": "TC025",
        "Category": "Output Validation",
        "Test Case Description": "Verify portfolio allocation sums to 100%",
        "Test Type": "Functional",
        "Preconditions": "Valid portfolio generation completed",
        "Test Steps": "1. Generate portfolio\n2. Extract all position allocations\n3. Sum percentages\n4. Verify total = 100% (±1% tolerance)",
        "Test Data": "Use TC006 request",
        "Expected Output": "portfolio.holdings array where:\nSum of all allocation_percent = 99-101%\nOR sum of position values ≈ capital",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Critical - allocation integrity"
    })
    
    test_cases.append({
        "Test Case #": "TC026",
        "Category": "Output Validation",
        "Test Case Description": "Verify performance_report contains key metrics",
        "Test Type": "Functional",
        "Preconditions": "Valid portfolio generation completed",
        "Test Steps": "1. Generate portfolio\n2. Inspect performance_report\n3. Verify metrics present: Sharpe ratio, volatility, beta, expected return",
        "Test Data": "Use TC006 request",
        "Expected Output": "performance_report contains:\n- expected_return (percentage)\n- volatility or std_dev\n- sharpe_ratio (number)\n- beta (if calculated)\n- max_drawdown (optional)",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Performance agent output"
    })
    
    test_cases.append({
        "Test Case #": "TC027",
        "Category": "Output Validation",
        "Test Case Description": "Verify backtest_results validates portfolio",
        "Test Type": "Functional",
        "Preconditions": "Valid portfolio generation completed",
        "Test Steps": "1. Generate portfolio\n2. Inspect backtest_results\n3. Verify contains historical returns over 12 months",
        "Test Data": "Use TC006 request",
        "Expected Output": "backtest_results contains:\n- period (e.g., 12 months)\n- cumulative_return\n- annualized_return\n- max_drawdown\n- win_rate or success metrics",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Backtest agent validation"
    })
    
    test_cases.append({
        "Test Case #": "TC028",
        "Category": "Output Validation",
        "Test Case Description": "Verify final_report is comprehensive markdown",
        "Test Type": "Functional",
        "Preconditions": "Valid portfolio generation completed",
        "Test Steps": "1. Generate portfolio\n2. Inspect final_report field\n3. Verify it's markdown with all sections",
        "Test Data": "Use TC006 request",
        "Expected Output": "final_report is markdown string containing:\n- Executive summary\n- Risk profile\n- Portfolio holdings table\n- Performance metrics\n- Investment rationale\n- Length > 500 characters",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Report synthesizer output"
    })
    
    # ==================== CATEGORY 6: API ERROR HANDLING ====================
    
    test_cases.append({
        "Test Case #": "TC029",
        "Category": "Error Handling",
        "Test Case Description": "Handle missing GOOGLE_API_KEY gracefully",
        "Test Type": "Negative",
        "Preconditions": "Remove or unset GOOGLE_API_KEY from .env",
        "Test Steps": "1. Start server without GOOGLE_API_KEY\n2. Send portfolio generation request\n3. Verify appropriate error message",
        "Test Data": "Use TC006 request",
        "Expected Output": "success=false\nerror message indicating API key missing or authentication failed",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Critical API key validation"
    })
    
    test_cases.append({
        "Test Case #": "TC030",
        "Category": "Error Handling",
        "Test Case Description": "Handle missing ALPHA_VANTAGE_API_KEY",
        "Test Type": "Negative",
        "Preconditions": "Remove ALPHA_VANTAGE_API_KEY from .env",
        "Test Steps": "1. Start server\n2. Send portfolio request\n3. Verify error indicates stock API issue",
        "Test Data": "Use TC006 request",
        "Expected Output": "success=false\nerror mentions Alpha Vantage or stock data unavailable",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Stock API dependency"
    })
    
    test_cases.append({
        "Test Case #": "TC031",
        "Category": "Error Handling",
        "Test Case Description": "Handle Alpha Vantage rate limit exceeded",
        "Test Type": "Negative",
        "Preconditions": "Exceed 5 API calls per minute",
        "Test Steps": "1. Make multiple rapid portfolio requests\n2. Observe behavior when rate limited\n3. Verify graceful degradation or retry logic",
        "Test Data": "Make 3-4 rapid requests",
        "Expected Output": "Either:\n- Requests queued and processed slowly\n- Error message about rate limiting\n- Backoff and retry mechanism",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Important for production use"
    })
    
    test_cases.append({
        "Test Case #": "TC032",
        "Category": "Error Handling",
        "Test Case Description": "Handle network timeout gracefully",
        "Test Type": "Negative",
        "Preconditions": "Simulate network issues or disconnect internet",
        "Test Steps": "1. Disconnect network or use proxy to block API calls\n2. Send portfolio request\n3. Verify timeout error handling",
        "Test Data": "Use TC006 request",
        "Expected Output": "success=false\nerror message about network connectivity or timeout\nNo server crash",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Network resilience test"
    })
    
    test_cases.append({
        "Test Case #": "TC033",
        "Category": "Error Handling",
        "Test Case Description": "Handle invalid stock symbol from agent",
        "Test Type": "Negative",
        "Preconditions": "Server running normally",
        "Test Steps": "1. Generate portfolio\n2. If agent picks invalid stock symbols, verify handling\n3. Check that portfolio generation doesn't crash",
        "Test Data": "Use TC006 request",
        "Expected Output": "Either:\n- Agent picks only valid symbols\n- Invalid symbols skipped with warning\n- Graceful error with partial results",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Agent output validation"
    })
    
    # ==================== CATEGORY 7: A2A PROTOCOL TESTING ====================
    
    test_cases.append({
        "Test Case #": "TC034",
        "Category": "A2A Protocol",
        "Test Case Description": "Send message via A2A /message:send endpoint",
        "Test Type": "Functional",
        "Preconditions": "Server running",
        "Test Steps": "1. Send POST to /message:send with proper A2A format\n2. Include message with text about creating portfolio\n3. Verify Task object returned",
        "Test Data": '{"message": {"role": "ROLE_USER", "parts": [{"text": "I want to invest $10,000"}]}, "contextId": "test-001"}',
        "Expected Output": "Status: 200 OK\nResponse contains task object with:\n- id\n- status.state = input_required or in_progress\n- status.message with agent response",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "A2A protocol compatibility"
    })
    
    test_cases.append({
        "Test Case #": "TC035",
        "Category": "A2A Protocol",
        "Test Case Description": "Multi-turn conversation via A2A",
        "Test Type": "Functional",
        "Preconditions": "Server running, TC034 completed",
        "Test Steps": "1. Send first message\n2. Use same contextId for follow-up message\n3. Verify conversation context maintained",
        "Test Data": 'Message 1: "I want to invest $15,000"\nMessage 2: "My goal is balanced growth" (same contextId)',
        "Expected Output": "Agent remembers previous messages\nResponses build on conversation history\nContext maintained across requests",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Stateful conversation test"
    })
    
    test_cases.append({
        "Test Case #": "TC036",
        "Category": "A2A Protocol",
        "Test Case Description": "Complete portfolio generation via A2A multi-turn",
        "Test Type": "Integration",
        "Preconditions": "Server running",
        "Test Steps": "1. Start conversation\n2. Provide all profile info across multiple messages\n3. Confirm portfolio generation\n4. Verify artifacts in final response",
        "Test Data": "Series of messages providing capital, goals, risk tolerance, etc.",
        "Expected Output": "Final Task with:\n- status.state = completed\n- artifacts array with portfolio JSON\n- comprehensive agent response",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "End-to-end A2A flow"
    })
    
    # ==================== CATEGORY 8: PERFORMANCE & TIMING ====================
    
    test_cases.append({
        "Test Case #": "TC037",
        "Category": "Performance",
        "Test Case Description": "Measure portfolio generation response time",
        "Test Type": "Performance",
        "Preconditions": "Server running with valid API keys",
        "Test Steps": "1. Record start time\n2. Send portfolio generation request\n3. Record end time when response received\n4. Calculate duration",
        "Test Data": "Use TC006 request",
        "Expected Output": "Response time: 30-120 seconds (due to sequential agents and API rate limits)\nAcceptable range depends on network and API speed",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Performance benchmark"
    })
    
    test_cases.append({
        "Test Case #": "TC038",
        "Category": "Performance",
        "Test Case Description": "Verify server handles concurrent requests",
        "Test Type": "Performance",
        "Preconditions": "Server running",
        "Test Steps": "1. Send 3 portfolio requests simultaneously\n2. Verify all complete successfully\n3. Check for any timeouts or failures",
        "Test Data": "3 different portfolio profiles",
        "Expected Output": "All 3 requests complete successfully\nMay take longer due to rate limiting\nNo server crashes or errors",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Concurrency test"
    })
    
    test_cases.append({
        "Test Case #": "TC039",
        "Category": "Performance",
        "Test Case Description": "Verify health endpoint response time",
        "Test Type": "Performance",
        "Preconditions": "Server running",
        "Test Steps": "1. Measure response time for /health endpoint\n2. Should be < 100ms",
        "Test Data": "N/A",
        "Expected Output": "Response time < 100ms\nImmediate response",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Health check should be fast"
    })
    
    # ==================== CATEGORY 9: DATA VALIDATION ====================
    
    test_cases.append({
        "Test Case #": "TC040",
        "Category": "Data Validation",
        "Test Case Description": "Verify all stock symbols are valid ticker formats",
        "Test Type": "Functional",
        "Preconditions": "Portfolio generated successfully",
        "Test Steps": "1. Generate portfolio\n2. Extract all stock symbols\n3. Verify format: 1-5 uppercase letters (US stocks)",
        "Test Data": "Use TC006 request",
        "Expected Output": "All ticker symbols match pattern: ^[A-Z]{1,5}$\nNo invalid characters or formats",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Data quality check"
    })
    
    test_cases.append({
        "Test Case #": "TC041",
        "Category": "Data Validation",
        "Test Case Description": "Verify all percentages are valid (0-100)",
        "Test Type": "Functional",
        "Preconditions": "Portfolio generated successfully",
        "Test Steps": "1. Generate portfolio\n2. Check all allocation_percent values\n3. Verify each is 0-100",
        "Test Data": "Use TC006 request",
        "Expected Output": "All allocation percentages: 0 ≤ value ≤ 100\nNo negative or >100 values",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Data boundary validation"
    })
    
    test_cases.append({
        "Test Case #": "TC042",
        "Category": "Data Validation",
        "Test Case Description": "Verify Sharpe ratio is reasonable",
        "Test Type": "Functional",
        "Preconditions": "Portfolio with performance_report generated",
        "Test Steps": "1. Generate portfolio\n2. Extract Sharpe ratio\n3. Verify it's in reasonable range (-2 to 4 for typical portfolios)",
        "Test Data": "Use TC006 request",
        "Expected Output": "Sharpe ratio approximately -2 to 4\n(Higher is better, >1 is good, >2 is excellent)",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Financial metric validation"
    })
    
    test_cases.append({
        "Test Case #": "TC043",
        "Category": "Data Validation",
        "Test Case Description": "Verify portfolio diversification across sectors",
        "Test Type": "Functional",
        "Preconditions": "Portfolio generated with moderate+ risk",
        "Test Steps": "1. Generate moderate risk portfolio\n2. Count unique sectors in holdings\n3. Verify at least 3-4 different sectors",
        "Test Data": "Use TC006 request (moderate profile)",
        "Expected Output": "Portfolio contains stocks from 3+ sectors\nNo single sector > 40% of portfolio\nProper diversification",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Diversification principle"
    })
    
    # ==================== CATEGORY 10: CROSS-PLATFORM COMPATIBILITY ====================
    
    test_cases.append({
        "Test Case #": "TC044",
        "Category": "Cross-Platform",
        "Test Case Description": "Verify UI loads correctly on Windows",
        "Test Type": "Compatibility",
        "Preconditions": "Windows system with Python 3.10+, Server running",
        "Test Steps": "1. Start server on Windows\n2. Open http://localhost:8000/ in browser\n3. Verify no UnicodeDecodeError\n4. Verify UI loads properly",
        "Test Data": "N/A",
        "Expected Output": "Status: 200 OK\nUI loads without encoding errors\nUTF-8 encoding properly handled",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Critical Windows compatibility (TC004 fix verification)"
    })
    
    test_cases.append({
        "Test Case #": "TC045",
        "Category": "Cross-Platform",
        "Test Case Description": "Verify portfolio generation on Windows",
        "Test Type": "Compatibility",
        "Preconditions": "Windows system, valid API keys, server running",
        "Test Steps": "1. Send portfolio generation request on Windows\n2. Verify full pipeline completes\n3. Check all outputs properly formatted",
        "Test Data": "Use TC006 request",
        "Expected Output": "Same behavior as macOS/Linux\nNo path or encoding issues\nPortfolio generated successfully",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Windows-specific testing"
    })
    
    test_cases.append({
        "Test Case #": "TC046",
        "Category": "Cross-Platform",
        "Test Case Description": "Verify portfolio generation on macOS",
        "Test Type": "Compatibility",
        "Preconditions": "macOS system, valid API keys, server running",
        "Test Steps": "1. Run complete portfolio generation\n2. Verify all agents execute properly\n3. Check outputs",
        "Test Data": "Use TC006 request",
        "Expected Output": "Portfolio generated successfully\nAll agents complete without errors\nProper JSON and markdown outputs",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "macOS verification"
    })
    
    test_cases.append({
        "Test Case #": "TC047",
        "Category": "Cross-Platform",
        "Test Case Description": "Verify portfolio generation on Linux",
        "Test Type": "Compatibility",
        "Preconditions": "Linux system, valid API keys, server running",
        "Test Steps": "1. Run complete portfolio generation\n2. Verify all functionality\n3. Check for any Linux-specific issues",
        "Test Data": "Use TC006 request",
        "Expected Output": "Portfolio generated successfully\nNo platform-specific issues\nSame results as other platforms",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Linux verification"
    })
    
    # ==================== CATEGORY 11: EDGE CASES & SPECIAL SCENARIOS ====================
    
    test_cases.append({
        "Test Case #": "TC048",
        "Category": "Edge Cases",
        "Test Case Description": "Handle profile with conflicting risk indicators",
        "Test Type": "Edge Case",
        "Preconditions": "Server running",
        "Test Steps": "1. Send request with conservative goal but high risk tolerance\n2. Verify agent handles conflict appropriately\n3. Check risk score calculation",
        "Test Data": '{"capital": 20000, "goal": "preserve_capital", "horizon": "less_than_1_year", "max_loss": 30, "experience": "expert", "income_stability": "stable"}',
        "Expected Output": "Risk score calculated as average (moderate)\nAgent may note the inconsistency\nPortfolio balances conflicting signals",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Conflicting input handling"
    })
    
    test_cases.append({
        "Test Case #": "TC049",
        "Category": "Edge Cases",
        "Test Case Description": "Generate portfolio during market hours vs after hours",
        "Test Type": "Edge Case",
        "Preconditions": "Server running at different times",
        "Test Steps": "1. Generate portfolio during US market hours (9:30am-4pm ET)\n2. Generate same portfolio after hours\n3. Compare results",
        "Test Data": "Use TC006 request at different times",
        "Expected Output": "Both generate successfully\nReal-time prices during market hours\nLast closing prices after hours\nMinor price differences but similar portfolios",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Market timing test"
    })
    
    test_cases.append({
        "Test Case #": "TC050",
        "Category": "Edge Cases",
        "Test Case Description": "Generate portfolio on weekend",
        "Test Type": "Edge Case",
        "Preconditions": "Server running on Saturday or Sunday",
        "Test Steps": "1. Generate portfolio when markets closed\n2. Verify uses last available data\n3. Check agent handles appropriately",
        "Test Data": "Use TC006 request on weekend",
        "Expected Output": "Portfolio generates successfully\nUses Friday's closing prices\nAgent may note markets are closed\nNo errors due to timing",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Weekend/holiday handling"
    })
    
    test_cases.append({
        "Test Case #": "TC051",
        "Category": "Edge Cases",
        "Test Case Description": "Verify behavior with very unstable income",
        "Test Type": "Edge Case",
        "Preconditions": "Server running",
        "Test Steps": "1. Send request with income_stability='very_unstable'\n2. Verify more conservative allocation\n3. Check risk score adjustment",
        "Test Data": '{"capital": 15000, "goal": "balanced_growth", "horizon": "3_5_years", "max_loss": 15, "experience": "intermediate", "income_stability": "very_unstable"}',
        "Expected Output": "Risk score slightly lower than stable income\nMore conservative allocation\nAgent may recommend larger emergency fund",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Income stability impact"
    })
    
    test_cases.append({
        "Test Case #": "TC052",
        "Category": "Edge Cases",
        "Test Case Description": "Generate multiple portfolios for same user",
        "Test Type": "Edge Case",
        "Preconditions": "Server running",
        "Test Steps": "1. Generate portfolio with specific profile\n2. Immediately generate another with same profile\n3. Compare results",
        "Test Data": "Use TC006 request twice",
        "Expected Output": "Both portfolios generate successfully\nMay have slight variations due to:\n- Real-time price changes\n- Agent reasoning differences\nOverall strategy should be similar",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Consistency check"
    })
    
    # ==================== CATEGORY 12: DOCUMENTATION & API SPECS ====================
    
    test_cases.append({
        "Test Case #": "TC053",
        "Category": "Documentation",
        "Test Case Description": "Verify /docs (Swagger UI) is accessible",
        "Test Type": "Functional",
        "Preconditions": "Server running",
        "Test Steps": "1. Navigate to http://localhost:8000/docs\n2. Verify Swagger UI loads\n3. Check all endpoints documented",
        "Test Data": "N/A",
        "Expected Output": "Swagger UI loads successfully\nAll endpoints listed:\n- /health\n- /api/info\n- /api/generate-portfolio\n- /message:send\n- /.well-known/agent-card.json\nRequest/response models documented",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "API documentation availability"
    })
    
    test_cases.append({
        "Test Case #": "TC054",
        "Category": "Documentation",
        "Test Case Description": "Verify /redoc (ReDoc) is accessible",
        "Test Type": "Functional",
        "Preconditions": "Server running",
        "Test Steps": "1. Navigate to http://localhost:8000/redoc\n2. Verify ReDoc loads\n3. Check documentation quality",
        "Test Data": "N/A",
        "Expected Output": "ReDoc UI loads successfully\nClear endpoint documentation\nRequest/response examples",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Alternative API documentation"
    })
    
    test_cases.append({
        "Test Case #": "TC055",
        "Category": "Documentation",
        "Test Case Description": "Try example request from Swagger UI",
        "Test Type": "Integration",
        "Preconditions": "Server running, /docs accessible",
        "Test Steps": "1. Open /docs\n2. Expand /api/generate-portfolio endpoint\n3. Click 'Try it out'\n4. Use example values\n5. Execute request",
        "Test Data": "Use Swagger UI default example",
        "Expected Output": "Request executes from Swagger UI\nResponse displayed in UI\nSuccess response with portfolio data",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Interactive documentation test"
    })
    
    # ==================== CATEGORY 13: SECURITY & VALIDATION ====================
    
    test_cases.append({
        "Test Case #": "TC056",
        "Category": "Security",
        "Test Case Description": "Verify API keys not exposed in responses",
        "Test Type": "Security",
        "Preconditions": "Server running with API keys configured",
        "Test Steps": "1. Generate portfolio\n2. Inspect all response fields\n3. Verify no API keys in response",
        "Test Data": "Use TC006 request",
        "Expected Output": "Response does NOT contain:\n- GOOGLE_API_KEY\n- ALPHA_VANTAGE_API_KEY\n- Any other sensitive credentials",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Critical security check"
    })
    
    test_cases.append({
        "Test Case #": "TC057",
        "Category": "Security",
        "Test Case Description": "Test SQL injection in capital field",
        "Test Type": "Security",
        "Preconditions": "Server running",
        "Test Steps": "1. Send request with SQL injection in capital field\n2. Verify rejected or harmless",
        "Test Data": '{"capital": "10000; DROP TABLE users;", "goal": "balanced_growth", "horizon": "3_5_years", "max_loss": 15, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "Request rejected with validation error\nOR harmless (pydantic converts to float, injection fails)\nNo database operations affected",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "SQL injection prevention"
    })
    
    test_cases.append({
        "Test Case #": "TC058",
        "Category": "Security",
        "Test Case Description": "Test XSS in text fields",
        "Test Type": "Security",
        "Preconditions": "Server running",
        "Test Steps": "1. Send request with XSS payload in goal field\n2. Verify sanitized or rejected",
        "Test Data": '{"capital": 10000, "goal": "<script>alert(XSS)</script>", "horizon": "3_5_years", "max_loss": 15, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "Request rejected (invalid goal value)\nOR payload sanitized/escaped\nNo script execution",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "XSS prevention"
    })
    
    test_cases.append({
        "Test Case #": "TC059",
        "Category": "Security",
        "Test Case Description": "Verify CORS headers are appropriate",
        "Test Type": "Security",
        "Preconditions": "Server running",
        "Test Steps": "1. Send request from browser/Postman\n2. Check CORS headers in response\n3. Verify Access-Control-Allow-Origin",
        "Test Data": "N/A",
        "Expected Output": "CORS headers present\nFor development: allow_origins=['*']\nFor production: should be specific domains",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "CORS configuration check"
    })
    
    test_cases.append({
        "Test Case #": "TC060",
        "Category": "Security",
        "Test Case Description": "Test extremely large capital value",
        "Test Type": "Security",
        "Preconditions": "Server running",
        "Test Steps": "1. Send request with very large capital (e.g., 999999999999)\n2. Verify handling\n3. Check for overflow issues",
        "Test Data": '{"capital": 999999999999, "goal": "balanced_growth", "horizon": "3_5_years", "max_loss": 15, "experience": "intermediate", "income_stability": "stable"}',
        "Expected Output": "Either:\n- Request accepted and portfolio generated\n- Validation error if exceeds reasonable limit\nNo system crash or overflow",
        "Actual Output": "",
        "Pass/Fail": "",
        "Comments": "Numeric overflow prevention"
    })
    
    return test_cases


def create_excel_file():
    """Create Excel file with test cases."""
    print("Generating Portfolio Manager Test Cases...")
    
    test_cases = generate_test_cases()
    df = pd.DataFrame(test_cases)
    
    # Create Excel file with formatting
    excel_path = "/Users/procal/Developer/Drayvn/code/drayvn_agents/portfolio_manager/Portfolio_Manager_Test_Cases.xlsx"
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Test Cases', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Test Cases']
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 12  # Test Case #
        worksheet.column_dimensions['B'].width = 18  # Category
        worksheet.column_dimensions['C'].width = 50  # Description
        worksheet.column_dimensions['D'].width = 15  # Test Type
        worksheet.column_dimensions['E'].width = 30  # Preconditions
        worksheet.column_dimensions['F'].width = 50  # Test Steps
        worksheet.column_dimensions['G'].width = 50  # Test Data
        worksheet.column_dimensions['H'].width = 50  # Expected Output
        worksheet.column_dimensions['I'].width = 50  # Actual Output
        worksheet.column_dimensions['J'].width = 12  # Pass/Fail
        worksheet.column_dimensions['K'].width = 30  # Comments
        
        # Format header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
    
    print(f"✓ Test cases generated successfully!")
    print(f"✓ File saved to: {excel_path}")
    print(f"✓ Total test cases: {len(test_cases)}")
    
    # Print summary by category
    print("\n=== Test Cases Summary by Category ===")
    category_counts = df['Category'].value_counts()
    for category, count in category_counts.items():
        print(f"  {category}: {count} test cases")
    
    return excel_path


if __name__ == "__main__":
    create_excel_file()
